from rest_framework.views import APIView
from django.http import HttpResponse
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill
)
from openpyxl.utils import get_column_letter

from datetime import timedelta
from django.utils.timezone import now

from activities.models import Activity
from rest_framework.permissions import IsAuthenticated
from clients.models import Client


class ExportExcelAPI(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        qs = Activity.objects.select_related(
            "user",
            "project"
        ).order_by("-date")

        export_type = (
            request.GET.get("export_type")
            or request.GET.get("type")
        )

        is_client_user = False

        try:

            client = Client.objects.get(
                user=request.user
            )

            is_client_user = True

            qs = qs.filter(
                project__client=client
            )

        except Client.DoesNotExist:

            if export_type == "daily":

                qs = qs.filter(
                    user=request.user
                )

            elif export_type == "approval":

                if not (
                    request.user.is_superuser
                    or request.user.is_staff
                ):

                    qs = qs.filter(
                        user=request.user
                    )

        project = request.GET.get("project")
        service = request.GET.get("service")
        task = request.GET.get("task")

        filter_type = (
            request.GET.get("date_filter")
            or request.GET.get("type")
            or request.GET.get("filter")
        )

        start = (
            request.GET.get("start_date")
            or request.GET.get("start")
        )

        end = (
            request.GET.get("end_date")
            or request.GET.get("end")
        )

        status = request.GET.get("status")
        search = request.GET.get("search")

        if project:

            qs = qs.filter(project_id=project)

        if service:

            qs = qs.filter(service_name=service)

        if task:

            qs = qs.filter(task_type=task)

        if status:

            qs = qs.filter(status=status)

        if search:

            qs = qs.filter(
                Q(user__first_name__icontains=search)
                |
                Q(user__last_name__icontains=search)
                |
                Q(project__name__icontains=search)
                |
                Q(task_type__icontains=search)
                |
                Q(service_name__icontains=search)
                |
                Q(category__icontains=search)
            )

        today = now().date()

        if filter_type == "today":

            qs = qs.filter(date=today)

        elif filter_type == "month":

            qs = qs.filter(
                date__month=today.month,
                date__year=today.year
            )

        elif filter_type == "week":

            start_week = (
                today - timedelta(
                    days=today.weekday()
                )
            )

            end_week = (
                start_week + timedelta(days=6)
            )

            qs = qs.filter(
                date__range=[
                    start_week,
                    end_week
                ]
            )

        elif filter_type == "year":

            qs = qs.filter(
                date__year=today.year
            )

        elif filter_type == "custom":

            if start and end:

                qs = qs.filter(
                    date__range=[start, end]
                )

        if start and end:

            qs = qs.filter(
                date__range=[start, end]
            )

        wb = Workbook()

        ws = wb.active

        ws.title = "SEO Report"

        if is_client_user:

            headers = [
                "S.No",
                "Date",
                "Service",
                "Module",
                "Task",
                "Keyword",
                "Target URL",
                "Submitted URL",
                "Other Data",
                "Hours",
            ]

        else:

            headers = [
                "S.No",
                "Date",
                "Employee",
                "Project",
                "Category",
                "Service",
                "Module",
                "Task",
                "Keyword",
                "Target URL",
                "Submitted URL",
                "Other Data",
                "Hours",
                "Status",
            ]

        ws.append(headers)

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for i, a in enumerate(qs, start=1):

            data = a.dynamic_data or {}

            employee = ""

            module_name = (
                data.get("module")
                or data.get("Module")
                or ""
            )

            hours = (
                data.get("hours")
                or data.get("Hours")
                or ""
            )

            if a.user:

                employee = (
                    f"{a.user.first_name} "
                    f"{a.user.last_name}"
                ).strip()

                if not employee:

                    employee = a.user.email

            keyword = (
                data.get("keyword")
                or data.get("KEYWORD")
                or data.get("Keyword")
                or ""
            )

            submitted_urls = (
                data.get("submitted_url")
                or data.get("submitted_urls")
                or data.get("SUBMITTED_URL")
                or data.get("Submitted_url")
                or ""
            )

            target_urls = (
                data.get("target_url")
                or data.get("target_urls")
                or data.get("TARGET_URL")
                or data.get("Target_url")
                or ""
            )

            submitted_list = [
                x.strip()
                for x in str(submitted_urls)
                    .replace(",", "\n")
                    .splitlines()
                if x.strip()
            ]

            if not submitted_list:

                submitted_list = [""]

            other_data_parts = []

            for key, value in data.items():

                lower_key = str(key).lower()

                if lower_key in [
                    "keyword",
                    "submitted_url",
                    "submitted_urls",
                    "target_url",
                    "target_urls",
                    "module",
                    "hours"
                ]:
                    continue

                if value is None or value == "":
                    continue

                if isinstance(value, list):

                    value = ", ".join(
                        map(str, value)
                    )

                other_data_parts.append(
                    f"{str(key).replace('_', ' ').title()}: {value}"
                )

            other_data = "\n".join(
                other_data_parts
            )

            start_row = ws.max_row + 1

            for idx, link in enumerate(submitted_list):

                if is_client_user:

                    # Columns: S.No | Date | Service | Module | Task |
                    #          Keyword | Target URL | Submitted URL |
                    #          Other Data | Hours
                    row = [

                        i if idx == 0 else "",

                        str(a.date)
                        if idx == 0 else "",

                        a.service_name or ""
                        if idx == 0 else "",

                        module_name
                        if idx == 0 else "",

                        a.task_type or ""
                        if idx == 0 else "",

                        keyword
                        if idx == 0 else "",

                        target_urls
                        if idx == 0 else "",

                        link,

                        other_data
                        if idx == 0 else "",

                        hours
                        if idx == 0 else "",
                    ]

                else:

                    # Columns: S.No | Date | Employee | Project | Category |
                    #          Service | Module | Task | Keyword | Target URL |
                    #          Submitted URL | Other Data | Hours | Status
                    row = [

                        i if idx == 0 else "",

                        str(a.date)
                        if idx == 0 else "",

                        employee
                        if idx == 0 else "",

                        
                        getattr(a.project, "name", "")
                        if idx == 0 else "",

                        getattr(a, "category", "")
                        if idx == 0 else "",

                        a.service_name or ""
                        if idx == 0 else "",

                        module_name
                        if idx == 0 else "",

                        a.task_type or ""
                        if idx == 0 else "",

                        keyword
                        if idx == 0 else "",

                        target_urls
                        if idx == 0 else "",

                        link,

                        other_data
                        if idx == 0 else "",

                        hours
                        if idx == 0 else "",

                        a.status.title()
                        if idx == 0 and a.status
                        else "",
                    ]

                ws.append(row)

                current_row = ws.max_row

                # Submitted URL column
                # Client:  col 8 (Submitted URL is 8th column)
                # Admin:   col 11 (Submitted URL is 11th column)
                submitted_col = 8 if is_client_user else 11

                submitted_cell = ws.cell(
                    current_row,
                    submitted_col
                )

                if link.startswith(
                    ("http://", "https://")
                ):

                    submitted_cell.hyperlink = link

                    submitted_cell.font = Font(
                        color="0000FF",
                        underline="single"
                    )

                # Target URL column
                # Client:  col 7 (Target URL is 7th column)
                # Admin:   col 10 (Target URL is 10th column)
                target_col = 7 if is_client_user else 10

                target_cell = ws.cell(
                    current_row,
                    target_col
                )

                if (
                    idx == 0
                    and str(target_urls).startswith(
                        ("http://", "https://")
                    )
                ):

                    target_cell.hyperlink = str(target_urls)

                    target_cell.font = Font(
                        color="0000FF",
                        underline="single"
                    )

            end_row = ws.max_row

            if len(submitted_list) > 1:

                if is_client_user:

                    # Merge all columns except Submitted URL (col 8)
                    merge_columns = [
                        1, 2, 3, 4, 5,
                        6, 7, 9, 10
                    ]

                else:

                    # Merge all columns except Submitted URL (col 11)
                    merge_columns = [
                        1, 2, 3, 4, 5,
                        6, 7, 8, 9, 10,
                        12, 13, 14
                    ]

                for col in merge_columns:

                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=end_row,
                        end_column=col
                    )

                    merged_cell = ws.cell(
                        start_row,
                        col
                    )

                    merged_cell.alignment = Alignment(
                        vertical="center",
                        horizontal="left",
                        wrap_text=True
                    )

        for row in ws.iter_rows():

            for cell in row:

                # Submitted URL column alignment
                if (
                    cell.column == 8
                    and is_client_user
                ):

                    cell.alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="left"
                    )

                elif (
                    cell.column == 11
                    and not is_client_user
                ):

                    cell.alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="left"
                    )

                # Other Data column alignment
                elif (
                    cell.column == 9
                    and is_client_user
                ):

                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="left"
                    )

                elif (
                    cell.column == 12
                    and not is_client_user
                ):

                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="left"
                    )

                else:

                    cell.alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="left"
                    )

        if is_client_user:

            # A=S.No, B=Date, C=Service, D=Module, E=Task,
            # F=Keyword, G=Target URL, H=Submitted URL,
            # I=Other Data, J=Hours
            widths = {
                "A": 10,
                "B": 15,
                "C": 25,
                "D": 20,
                "E": 25,
                "F": 35,
                "G": 55,
                "H": 135,
                "I": 180,
                "J": 12,
            }

        else:

            # A=S.No, B=Date, C=Employee, D=Project, E=Category,
            # F=Service, G=Module, H=Task, I=Keyword, J=Target URL,
            # K=Submitted URL, L=Other Data, M=Hours, N=Status
            widths = {
                "A": 10,
                "B": 15,
                "C": 25,
                "D": 25,
                "E": 20,
                "F": 25,
                "G": 20,
                "H": 30,
                "I": 35,
                "J": 55,
                "K": 135,
                "L": 180,
                "M": 12,
                "N": 15,
            }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        for row in range(
            2,
            ws.max_row + 1
        ):

            other_col = 9 if is_client_user else 12

            other_value = ws.cell(
                row=row,
                column=other_col
            ).value

            submitted_col = 8 if is_client_user else 11

            submitted_value = ws.cell(
                row=row,
                column=submitted_col
            ).value

            # Default height
            current_height = 38

            # Submitted URL rows spacing
            if submitted_value:

                current_height = max(
                    current_height,
                    42
                )

            # Other Data auto height
            if other_value:

                text = str(other_value)

                estimated_lines = max(
                    text.count("\n") + 1,
                    len(text) // 85 + 1
                )

                current_height = max(
                    current_height,
                    min(estimated_lines * 24, 500)
                )

            ws.row_dimensions[row].height = current_height

        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        parts = ["Report"]

        if filter_type == "today":

            parts.append(
                now().strftime("%d-%m-%Y")
            )

        elif filter_type == "week":

            start_week = (
                today - timedelta(
                    days=today.weekday()
                )
            )

            end_week = (
                start_week + timedelta(days=6)
            )

            parts.append(
                f"{start_week}_to_{end_week}"
            )

        elif filter_type == "month":

            parts.append(
                now().strftime("%B-%Y")
            )

        elif filter_type == "year":

            parts.append(
                str(today.year)
            )

        elif filter_type == "custom":

            if start and end:

                parts.append(
                    f"{start}_to_{end}"
                )

        if service:

            parts.append(
                service.replace(" ", "_")
            )

        if task:

            parts.append(
                task.replace(" ", "_")
            )

        filename = (
            "_".join(parts)
            + ".xlsx"
        )

        response[
            "Content-Disposition"
        ] = (
            f'attachment; filename="{filename}"'
        )

        wb.save(response)

        return response
