from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count
from django.http import HttpResponse

from datetime import timedelta, datetime

from projects.models import Project
from activities.models import Activity
from accounts.models import Profile

from openpyxl import Workbook
from openpyxl.styles import Alignment
import openpyxl

@login_required
def activity_daily(request):

    projects = Project.objects.all().order_by('name')

    return render(request, 'frontend/activities/daily.html', {
        'projects': projects
    })

from django.db.models import Q
from django.db.models import Q, Count
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime

import openpyxl

from openpyxl.styles import (
    Alignment,
    Font,
    Border,
    Side,
    PatternFill
)

from activities.models import Activity
from projects.models import Project

from django.db.models import Q
from django.db.models.functions import Cast
from django.db.models import CharField

@login_required
def activity_approval(request):

    status = request.GET.get(
        'status',
        'pending'
    )

    date_filter = request.GET.get(
        'date',
        'all'
    )

    search = request.GET.get(
        'search',
        ''
    )

    project_id = request.GET.get(
        'project',
        ''
    )

    today = timezone.now().date()

    queryset = Activity.objects.select_related(
        'user',
        'project'
    )

    if date_filter == "today":

        queryset = queryset.filter(
            date=today
        )

    elif date_filter == "week":

        queryset = queryset.filter(
            date__gte=today - timedelta(days=6)
        )

    elif date_filter == "month":

        queryset = queryset.filter(
            date__month=today.month,
            date__year=today.year
        )

    elif date_filter == "year":

        queryset = queryset.filter(
            date__year=today.year
        )

    elif date_filter == "custom":

        selected_date = request.GET.get(
            "selected_date"
        )

        if selected_date:

            try:

                selected_date = datetime.strptime(
                    selected_date,
                    "%Y-%m-%d"
                ).date()

                queryset = queryset.filter(
                    date=selected_date
                )

            except:

                pass

    if search:

        search = search.strip()

        queryset = queryset.filter(

            Q(user__email__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(project__name__icontains=search) |
            Q(category__icontains=search) |
            Q(service_name__icontains=search) |
            Q(task_type__icontains=search) |

            Q(dynamic_data__keyword__icontains=search) |
            Q(dynamic_data__Keyword__icontains=search) |
            Q(dynamic_data__KEYWORD__icontains=search)

        )
    # PROJECT FILTER
    if project_id:

        queryset = queryset.filter(
            project_id=project_id
        )


    # STATUS COUNTS
    counts = queryset.values(
        'status'
    ).annotate(
        count=Count('id')
    )

    count_map = {
        i['status']: i['count']
        for i in counts
    }

    pending_count = count_map.get(
        'pending',
        0
    )

    approved_count = count_map.get(
        'approved',
        0
    )

    rejected_count = count_map.get(
        'rejected',
        0
    )
    
    if request.GET.get('export') == "excel":

        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "Activity Report"

   
        headers = [

            "S.No",

            "Project",

            "Employee",

            "Category",

            "Service",

            "Task",

            "Keyword",

            "Submitted URL",

            "Status",

            "Date"
        ]

        ws.append(headers)

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="4F46E5",
            fill_type="solid"
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        wrap = Alignment(
            wrap_text=True,
            vertical="top"
        )

        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        for cell in ws[1]:

            cell.font = header_font

            cell.fill = header_fill

            cell.alignment = center

            cell.border = border

    
        row_num = 2

        serial = 1

        for a in queryset.order_by('-created_at'):

            dynamic = a.dynamic_data or {}

            keyword = dynamic.get(
                "keyword",
                ""
            )

            submitted_url = (

                dynamic.get("submitted_url")

                or

                dynamic.get("url")

                or

                dynamic.get("proof_link")

                or

                ""
            )

            links = [

                l.strip()

                for l in str(
                    submitted_url
                ).split(",")

                if l.strip()

            ] or [""]

            start_row = row_num

            for i, link in enumerate(links):

                ws.cell(
                    row=row_num,
                    column=1,
                    value=serial if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=2,
                    value=a.project.name
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=3,
                    value=a.user.email
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=4,
                    value=a.category
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=5,
                    value=a.service_name
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=6,
                    value=a.task_type
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=7,
                    value=keyword
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=8,
                    value=link
                )

                ws.cell(
                    row=row_num,
                    column=9,
                    value=a.status
                    if i == 0 else ""
                )

                ws.cell(
                    row=row_num,
                    column=10,
                    value=str(a.date)
                    if i == 0 else ""
                )

                # HYPERLINK
                if link:

                    cell = ws.cell(
                        row=row_num,
                        column=8
                    )

                    cell.hyperlink = link

                    cell.style = "Hyperlink"

                # BORDER + ALIGNMENT
                for col in range(1, 11):

                    c = ws.cell(
                        row=row_num,
                        column=col
                    )

                    c.border = border

                    if col == 8:

                        c.alignment = wrap

                    else:

                        c.alignment = center

                row_num += 1

            end_row = row_num - 1

            if end_row > start_row:

                for col in [
                    1,
                    2,
                    3,
                    4,
                    5,
                    6,
                    7,
                    9,
                    10
                ]:

                    ws.merge_cells(

                        start_row=start_row,

                        start_column=col,

                        end_row=end_row,

                        end_column=col
                    )

   
            status_cell = ws.cell(
                row=start_row,
                column=9
            )

            if a.status == "approved":

                status_cell.fill = PatternFill(
                    start_color="D1FAE5",
                    fill_type="solid"
                )

            elif a.status == "pending":

                status_cell.fill = PatternFill(
                    start_color="FEF3C7",
                    fill_type="solid"
                )

            elif a.status == "rejected":

                status_cell.fill = PatternFill(
                    start_color="FEE2E2",
                    fill_type="solid"
                )

            serial += 1

    
        widths = [

            8,
            24,
            28,
            18,
            22,
            24,
            24,
            55,
            16,
            16
        ]

        for i, w in enumerate(
            widths,
            start=1
        ):

            ws.column_dimensions[
                chr(64 + i)
            ].width = w

  
        ws.auto_filter.ref = "A1:J1"

        ws.freeze_panes = "A2"

   
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response[
            'Content-Disposition'
        ] = 'attachment; filename=activity_report.xlsx'

        wb.save(response)

        return response

    activities = queryset.filter(
        status=status
    ).order_by('-created_at')

    projects = Project.objects.all()

    profile = Profile.objects.filter(
        user=request.user
    ).first()


    return render(

        request,

        'frontend/activities/approval.html',

        {

            'activities': activities,

            'projects': projects,

            'current_status': status,

            'date_filter': date_filter,

            'pending_count': pending_count,

            'approved_count': approved_count,

            'rejected_count': rejected_count,

            'search': search,

            'selected_project': project_id,

            'profile': profile
        }
    )


@login_required
def activity_reports(request):

    activities = Activity.objects.select_related(
        'user',
        'project'
    ).order_by('-date')

    return render(

        request,

        'frontend/activities/reports.html',

        {
            'activities': activities
        }
    )
